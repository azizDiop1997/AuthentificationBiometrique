	#https://www.tensorflow.org/tutorials/load_data/images
import matplotlib.pyplot as plt
import numpy as np
import os
import PIL
import tensorflow as tf
import glob
import cv2 
# in : 
# /home/kali/Bureau/PFE/projetGit/AuthentificationBiometrique
# python3 -m http.server  
# only 1st time to create the keras dataset locally (at /home/kali/.keras/datasets/flower_photos)
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras.models import Sequential
import openpyxl 

import pathlib

import argparse

from subprocess import *

img_height = 320
img_width = 120
ds_size=756
nb_classes=126


parser = argparse.ArgumentParser()

parser.add_argument('-e','--epochs',
required=True,
dest='epochs',
help='Epochs number, determines how long the algorithm will train')

parser.add_argument('-d','--demo',
dest='demo',
help='Displays learning graph of the model & validarion tests, False by default')

parser.add_argument('-v','--verbose',
dest='verb',
help='Displays some informations such as model\'s summary, number of classes. False by default')

parser.add_argument('-o','--order',
dest='orderImg',
help='Order the new images of /home/kali/Bureau/PFE/newGit/AuthentificationBiometrique/database/ directory, add them to the dataset')


args = parser.parse_args()


def setup(epochs):

	##dataset_url="http://0.0.0.0:8000/BDD_FingerVeins.tar.gz"
	##data_dir = tf.keras.utils.get_file(fname='DB_augmented', origin=dataset_url, untar=True) # or  just import data into /home/user/.keras/datasets
	#data_dir = pathlib.Path(data_dir)
	data_dir = pathlib.Path("/home/kali/.keras/datasets/DB_augmented") # extract_
	#image_count = len(list(data_dir.glob('*/*.bmp')))
	#print("images used in dataset: ",image_count)
	
	imgs = list(data_dir.glob('*/*.bmp')) # images dans les sous dossiers correspondant à leur classe
	
	batch_size = 32

	train_ds = tf.keras.utils.image_dataset_from_directory(
  data_dir,
  validation_split=0.2,
  subset="training",
  seed=123,
  image_size=(img_height, img_width),
  batch_size=batch_size)

	val_ds = tf.keras.utils.image_dataset_from_directory(
  data_dir,
  validation_split=0.2,
  subset="validation",
  seed=123,
  image_size=(img_height, img_width),
  batch_size=batch_size)


#	for image_batch, labels_batch in train_ds:
#		print(image_batch.shape)
#		print(labels_batch.shape)
#		break

	class_names = train_ds.class_names
	if args.verb:	
		print("Classes: ")
		print(class_names)
	
	#print(train_ds)
##	plt.figure(figsize=(10, 10))
#	for images, labels in train_ds.take(1):
#		for i in range(9):
#			ax = plt.subplot(3, 3, i + 1)
#			plt.imshow(images[i].numpy().astype("uint8"))
#			plt.title(class_names[labels[i]])
##			plt.axis("off")


#	for image_batch, labels_batch in train_ds:
#		print(image_batch.shape)
#		print(labels_batch.shape)
#		break

	
# https://www.tensorflow.org/tutorials/load_data/images#standardize_the_data

	
	# normalization
#	normalization_layer = tf.keras.layers.Rescaling(1./255)
	
#	normalized_ds = train_ds.map(lambda x, y: (normalization_layer(x), y))
	
#	print(normalized_ds)
#	print("------")	
#
#	image_batch, labels_batch = next(iter(normalized_ds))
#	print(image_batch, labels_batch )
#	first_image = image_batch[0]
	# Notice the pixel values are now in `[0,1]`.
#	print(np.min(first_image), np.max(first_image))


# 0.0 0.96902645


	model = tf.keras.Sequential([
  tf.keras.layers.Rescaling(1./255),
  tf.keras.layers.Conv2D(32, 3, activation='relu'),
  tf.keras.layers.MaxPooling2D(),
  tf.keras.layers.Conv2D(32, 3, activation='relu'),
  tf.keras.layers.MaxPooling2D(),
  tf.keras.layers.Conv2D(32, 3, activation='relu'),
  tf.keras.layers.MaxPooling2D(),
  tf.keras.layers.Dropout(0.2),
  tf.keras.layers.Flatten(),
  tf.keras.layers.Dense(128, activation='relu'),
  tf.keras.layers.Dense(nb_classes) # nb couches sorties/classes
])


	model.compile(
  optimizer='adam',
  loss=tf.losses.SparseCategoricalCrossentropy(from_logits=True),
  metrics=['accuracy'])

	history = model.fit(train_ds,validation_data=val_ds,epochs=epochs)

	if args.verb:
		model.summary()

	test_images=val_ds
	probability_model = tf.keras.Sequential([model, tf.keras.layers.Softmax()])
	predictions = probability_model.predict(test_images)

	#for b in range(5):
	##for images, labels in test_images.take(1):
		#for i in range(1):
				#ax = plt.subplot(3, 3, i + 1)
				#plt.imshow(images[i].numpy().astype("uint8"))
				#plt.title(class_names[labels[i]])
			#print(i,": ",class_names[labels[i]])
				#plt.axis("off")


	#print(predictions[0])
	#z=np.argmax(predictions[0])
	#print(z)
	#print(class_names[z])

	#print("-----------")
	#print(predictions[70])	
#	#print(len(predictions[150]))
	#y=np.argmax(predictions[70])
	#print(y)
	#print(class_names[y])

	#if args.demo:
	acc = history.history['accuracy']
	val_acc = history.history['val_accuracy']
	
	loss = history.history['loss']
	val_loss = history.history['val_loss']

	epochs_range = range(epochs)

	plt.figure(figsize=(8, 8))
	plt.subplot(1, 2, 1)
	plt.plot(epochs_range, acc, label='Training Accuracy')
	plt.plot(epochs_range, val_acc, label='Validation Accuracy')
	plt.legend(loc='lower right')
	plt.title('Training and Validation Accuracy')
	
	plt.subplot(1, 2, 2)
	plt.plot(epochs_range, loss, label='Training Loss')
	plt.plot(epochs_range, val_loss, label='Validation Loss')
	plt.legend(loc='upper right')
	plt.title('Training and Validation Loss')
	plt.show()
	
	# validation tests
	# if args.demo:
	p1 = Popen(["ls", "/home/kali/.keras/datasets/DB_augmented"], stdout=PIPE)
	p2 = Popen(["grep", "bmp"], stdin=p1.stdout, stdout=PIPE)
	output = p2.communicate()[0]
	x = (output.decode("utf-8")).split("\n")
	im_paths=x
	im_paths.pop(-1) # delete last element which is ""

	#print("!!!!!!!!")
	#print(im_paths)	
	#print("!!!!!!!!")
		
	for im in im_paths:
		im="/home/kali/.keras/datasets/DB_augmented/"+im
		print("- - - -")
		targ="null"
		for i in class_names:
			if str(i) in im:
				targ=str(i)
		print("target,image={} , {}".format(targ,im))
		
		img = tf.keras.utils.load_img(im, target_size=(img_height, img_width))
		img_array = tf.keras.utils.img_to_array(img)
		img_array = tf.expand_dims(img_array, 0) # Create a batch

		predictions = model.predict(img_array)
		score = tf.nn.softmax(predictions[0])

		res=class_names[np.argmax(score)]
		boolean="false"
		if str(targ)==str(res):
			boolean="true"
		if targ=="null":
			print("target doesn't exist.")
		else:		
			print("This image most likely belongs to {} with a {:.2f} percent confidence. => {}".format(class_names[np.argmax(score)], (100 * np.max(score)), boolean))

	return model, class_names





def orderNewImages():
	origin="/home/kali/Bureau/PFE/newGit/AuthentificationBiometrique/database/"
	images = glob.glob('{}*.bmp'.format(origin)) # contains images that have been added with data augmentation (file.py)
	
	for img in images: 
		#print(img)
		# traitement d'image & redimension.
		os.system('python3 /home/kali/Bureau/PFE/newGit/AuthentificationBiometrique/pycode_traitement_images/traitement_imagesv1.py -i {} -o {}'.format(img,img))

		short=img.split("/")[-1]
		#print("/",short)

		short=short.split(".")[0]
		#print(".",short)

		destClass=short.removeprefix("BDD_FingerVeins_original_")[:-2]
		#print("dC",destClass)
		dest="/home/kali/.keras/datasets/DB_augmented/"+destClass+"/"

		os.system("mv {} {}".format(img,dest))
		#print("moved:{} --TO-> {}".format(img,dest))


if args.orderImg:
	print("\t started ordering new images...")
	orderNewImages()
	print("\t images ready to join the dataset.")

setup(int(args.epochs))













# Realize nb_iter builds of the model, each with nb_epochs epochs. 
#Each build is striclty independant from the other, it permits to verify that the model generated is almost always the same
def excel_models(nb_iter,nb_epochs):
  
	# each model will test the following image
	impath = "/home/kali/.keras/datasets/004left_index_1.bmp"  
	targ=impath.split("/")[-1]
	targ=targ[:-6]
	
	target=[]
	result=[]
	confidence=[]
	correct=[]

	for i in range(nb_iter):
		img = tf.keras.utils.load_img(impath, target_size=(img_height, img_width))
		img_array = tf.keras.utils.img_to_array(img)
		img_array = tf.expand_dims(img_array, 0)
		m,classes=setup(nb_epochs)
		predictions = m.predict(img_array)
		score = tf.nn.softmax(predictions[0])
		res=classes[np.argmax(score)]
		boolean="false"
		if str(targ)==str(res):
			boolean="true"
			print("ok T")
		print("{} most likely belongs to {} class with a {:.2f} % confidence => {}".format(targ,res, 100 * np.max(score),boolean))
		target.append(targ)
		result.append(res)
		confidence.append("{:.2f}".format(100 * np.max(score)))
		correct.append(boolean)				

	# writing datas on excel
	wb = openpyxl.Workbook()  
	sheet = wb.active 

	ca = sheet.cell(row = 1, column = 1)
	ca.value = "Target class"
	cb = sheet.cell(row = 1, column = 2) 
	cb.value = "Predicted class"	
	cc = sheet.cell(row = 1, column = 3) 
	cc.value = "Confidence %"
	cd = sheet.cell(row = 1, column = 4) 
	cd.value = "Correctness"
	ce = sheet.cell(row = 1, column = 5)
	ce.value = "Iteration"
	for i in range(2,nb_iter+2):
		print("i: ",i)
		c1 = sheet.cell(row = i, column = 1) 
		c1.value = target[i-2]
		c2 = sheet.cell(row = i, column = 2) 
		c2.value = result[i-2]
		c3 = sheet.cell(row = i, column = 3) 
		s=confidence[i-2]#.replace(".",",")
		c3.value = float(s)
		c4 = sheet.cell(row = i, column = 4) 
		c4.value = correct[i-2]
		c5 = sheet.cell(row = i, column = 5) 
		c5.value = i-2	

	
	cmt = sheet.cell(row = nb_iter+2, column = 4)
	cmt.value = 'Average confidence'
	cm = sheet.cell(row = nb_iter+2, column = 5)	
	cm.value = '=AVERAGE(C2:C{})'.format(nb_iter+2)

	ctf = sheet.cell(row = nb_iter+3, column = 4)
	ctf.value = 'True/False ratio'
	tfr = sheet.cell(row = nb_iter+2, column = 4)
	for e in correct:
		if e=="true":
			ratio+=1
	tfr.value = '={}/{}'.format(ratio,nb_iter)


	wb.save("/home/kali/Bureau/PFE/out-{}models-{}epochs.xlsx".format(nb_iter,nb_epochs)) 


#excel_models(5,5)




